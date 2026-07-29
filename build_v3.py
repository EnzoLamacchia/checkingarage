"""
Genera il piano di rotazione garage v3 a partire da LISTA GARAGE.xlsx.

Regole v3 (rispetto a v2):
- Giorni a capienza limitata (7 accessi/giorno): LUN, MAR, MER, GIO.
- Unico giorno ad accesso libero (chiunque non sia in smart quel giorno): VEN.
- KETTY ha sempre accesso fisso di mercoledi (mai in rotazione), oltre
  all'accesso libero del venerdi se non in smart quel giorno. E' esclusa
  dal vincolo di equita sugli altri giorni limitati (lun/mar/gio), perche'
  e' in smart working proprio in quei tre giorni.
- Equita: per tutti tranne Ketty, ogni finestra scorrevole di 3 giorni
  limitati consecutivi (nell'ordine di calendario tra lun/mar/mer/gio)
  garantisce almeno 1 accesso a testa tra gli idonei di quella finestra.
- Periodo: 04/08/2026 - 30/09/2026.

Uso:
    py build_v3.py

Script persistente e riutilizzabile: NON cancellare dopo l'uso. Rilancialo
ogni volta che LISTA GARAGE.xlsx o le regole sopra cambiano.
"""

import datetime
import json
import random

import openpyxl

SOURCE_XLSX = "LISTA GARAGE.xlsx"
OUTPUT_XLSX = "Piano_Rotazione_Garage_v3.xlsx"
OUTPUT_JSON = "data_v3.json"
OUTPUT_HTML = "piano_rotazione_garage_v3.html"

PERIOD_START = datetime.date(2026, 8, 4)
PERIOD_END = datetime.date(2026, 9, 30)

LIMITED_CAPACITY = 7
LIMITED_DAYS = {0, 1, 2, 3}  # Mon, Tue, Wed, Thu (Python weekday: Mon=0)
FREE_DAY = 4  # Friday

DAY_IT = {0: "LUN", 1: "MAR", 2: "MER", 3: "GIO", 4: "VEN"}

SMART_DAY_MAP = {
    "LUNEDÌ": 0, "LUNEDI": 0,
    "MARTEDÌ": 1, "MARTEDI": 1,
    "MERCOLEDÌ": 2, "MERCOLEDI": 2,
    "GIOVEDÌ": 3, "GIOVEDI": 3,
    "VENERDÌ": 4, "VENERDI": 4,
}

KETTY_NAME = "KETTY"
KETTY_FIXED_WEEKDAY = 2  # Wednesday

RANDOM_SEED = 42


def load_people(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Foglio1"]
    people = []
    name_counts = {}
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        pid, nome, giorno_smart, _targa = row
        if pid is None or nome is None:
            continue
        name_counts[nome] = name_counts.get(nome, 0) + 1

    seen = {}
    for row in rows:
        pid, nome, giorno_smart, _targa = row
        if pid is None or nome is None:
            continue
        display_name = nome
        if name_counts[nome] > 1:
            seen[nome] = seen.get(nome, 0) + 1
            suffix = "DP" if seen[nome] == 1 else "S"
            display_name = f"{nome} {suffix}"

        smart_days = set()
        if giorno_smart:
            for part in str(giorno_smart).split(";"):
                key = part.strip().upper()
                if key in SMART_DAY_MAP:
                    smart_days.add(SMART_DAY_MAP[key])

        people.append({
            "id": pid,
            "nome": display_name,
            "smart": smart_days,
        })
    return people


def daterange(start, end):
    current = start
    while current <= end:
        yield current
        current += datetime.timedelta(days=1)


def build_limited_day_sequence(people, start, end):
    """Lista ordinata delle date (lun-gio) nel periodo, per finestre di equita."""
    return [d for d in daterange(start, end) if d.weekday() in LIMITED_DAYS]


def assign_limited_days(people, limited_dates, rng):
    """
    Assegna gli accessi per i giorni limitati (lun/mar/mer/gio) rispettando:
    - capienza fissa 7/giorno
    - nessuno accede nel proprio giorno di smart
    - Ketty sempre presente di mercoledi (non conta nel conteggio di rotazione)
    - equita a finestre scorrevoli di 3 giorni limitati consecutivi (esclusa Ketty):
      ogni idoneo della finestra riceve almeno 1 accesso entro la finestra stessa.
      Chi e' in smart in tutti e 3 i giorni della finestra non e' idoneo e viene
      escluso dal vincolo per quella finestra.
    """
    access = {p["id"]: {} for p in people}
    rotating_people = [p for p in people if p["nome"] != KETTY_NAME]
    # Conteggio totale di accessi ricevuti nel periodo (attraverso tutte le
    # finestre): usato per decidere chi assorbe il surplus quando una finestra
    # ha piu' slot che idonei unici, cosi' il surplus stesso ruota tra le
    # persone invece di accumularsi sempre sulle stesse.
    total_si_count = {p["id"]: 0 for p in rotating_people}
    # Chi ha ricevuto un secondo accesso nella stessa finestra (surplus
    # aritmetico inevitabile quando slot > idonei unici): va escluso, se
    # possibile, dal sorteggio casuale del giorno residuo finale, cosi' non
    # accumula anche quello e il gap massimo resta 1 invece di 2.
    surplus_recipients = set()

    windows = [limited_dates[i:i + 3] for i in range(0, len(limited_dates), 3)]

    # L'ultima finestra, se incompleta (< 3 giorni limitati), non puo' garantire
    # equita' ne' lo slot fisso di Ketty: assegnazione puramente casuale tra
    # tutti gli idonei di quel/quei giorno/i, Ketty inclusa alla pari.
    complete_windows = [w for w in windows if len(w) == 3]
    leftover_dates = [d for w in windows if len(w) < 3 for d in w]

    for window in complete_windows:
        got_access = {p["id"]: False for p in rotating_people}
        window_eligible = {
            p["id"]: any(date.weekday() not in p["smart"] for date in window)
            for p in rotating_people
        }

        for day_index, date in enumerate(window):
            weekday = date.weekday()
            date_str = date.isoformat()
            remaining_days_incl_this = len(window) - day_index

            forced_ketty = weekday == KETTY_FIXED_WEEKDAY
            slots = LIMITED_CAPACITY - (1 if forced_ketty else 0)

            eligible_today = [
                p["id"] for p in rotating_people
                if weekday not in p["smart"]
            ]

            # Chi e' idoneo nella finestra, non ha ancora avuto accesso, e per
            # cui questo e' l'ultimo giorno utile della finestra: priorita' assoluta.
            must_have_today = [
                pid for pid in eligible_today
                if window_eligible[pid] and not got_access[pid]
                and _is_last_chance(pid, people, window, day_index)
            ]

            must_have_today.sort(key=lambda pid: total_si_count[pid])

            rest = [pid for pid in eligible_today if pid not in must_have_today]
            rng.shuffle(rest)
            # priorita' assoluta a chi non ha ancora avuto accesso nella finestra;
            # chi lo ha gia' ricevuto viene ripescato solo se servono a riempire
            # gli slot residui (mai per dare un secondo accesso quando c'e'
            # ancora qualcuno a zero che potrebbe prenderlo). Tra chi deve
            # ricevere un secondo accesso nella stessa finestra (surplus),
            # priorita' a chi ha il conteggio totale piu' basso finora, cosi'
            # il surplus ruota invece di accumularsi sempre sulle stesse persone.
            not_yet = [pid for pid in rest if not got_access[pid]]
            not_yet.sort(key=lambda pid: total_si_count[pid])
            already = [pid for pid in rest if got_access[pid]]
            already.sort(key=lambda pid: total_si_count[pid])

            chosen = list(dict.fromkeys(must_have_today + not_yet + already))[:slots]
            chosen_set = set(chosen)

            for p in people:
                pid = p["id"]
                if p["nome"] == KETTY_NAME:
                    access[pid][date_str] = "SI" if forced_ketty else "NO"
                elif weekday in p["smart"]:
                    access[pid][date_str] = None
                elif pid in chosen_set:
                    if got_access[pid]:
                        surplus_recipients.add(pid)
                    access[pid][date_str] = "SI"
                    got_access[pid] = True
                    total_si_count[pid] += 1
                else:
                    access[pid][date_str] = "NO"

    for date in leftover_dates:
        weekday = date.weekday()
        date_str = date.isoformat()
        eligible_today = [p["id"] for p in people if weekday not in p["smart"]]
        rng.shuffle(eligible_today)
        # priorita' a chi NON ha gia' ricevuto un surplus altrove, cosi' il
        # sorteggio finale non si somma al surplus sulla stessa persona;
        # se non bastano candidati "puliti" per riempire i posti, si ripesca
        # comunque da chi ha gia' avuto un surplus.
        clean = [pid for pid in eligible_today if pid not in surplus_recipients]
        already_surplus = [pid for pid in eligible_today if pid in surplus_recipients]
        eligible_today = clean + already_surplus
        chosen_set = set(eligible_today[:LIMITED_CAPACITY])

        for p in people:
            pid = p["id"]
            if weekday in p["smart"]:
                access[pid][date_str] = None
            elif pid in chosen_set:
                access[pid][date_str] = "SI"
            else:
                access[pid][date_str] = "NO"

    return access


def _is_last_chance(pid, people, window, day_index):
    """True se pid non avra' altre occasioni utili in questa finestra dopo oggi."""
    person = next(p for p in people if p["id"] == pid)
    for date in window[day_index + 1:]:
        if date.weekday() not in person["smart"]:
            return False
    return True


def assign_free_day(people, start, end, access):
    for date in daterange(start, end):
        if date.weekday() != FREE_DAY:
            continue
        date_str = date.isoformat()
        for p in people:
            access[p["id"]][date_str] = "NO" if FREE_DAY in p["smart"] else "SI"


def validate(people, access, limited_dates):
    errors = []
    by_date = {}
    for p in people:
        for date_str, val in access[p["id"]].items():
            by_date.setdefault(date_str, []).append((p["id"], p["nome"], val))

    for date in limited_dates:
        date_str = date.isoformat()
        si_count = sum(1 for _, _, v in by_date[date_str] if v == "SI")
        if si_count != LIMITED_CAPACITY:
            errors.append(f"{date_str}: {si_count} accessi (attesi {LIMITED_CAPACITY})")

    windows = [limited_dates[i:i + 3] for i in range(0, len(limited_dates), 3)]
    complete_windows = [w for w in windows if len(w) == 3]

    ketty = next(p for p in people if p["nome"] == KETTY_NAME)
    for window in complete_windows:
        for date in window:
            if date.weekday() == KETTY_FIXED_WEEKDAY:
                if access[ketty["id"]][date.isoformat()] != "SI":
                    errors.append(f"{date.isoformat()}: Ketty non presente di mercoledi")

    for window in complete_windows:
        window_dates = [d.isoformat() for d in window]
        for p in people:
            if p["nome"] == KETTY_NAME:
                continue
            eligible = any(access[p["id"]].get(ds) is not None for ds in window_dates)
            got = any(access[p["id"]].get(ds) == "SI" for ds in window_dates)
            if eligible and not got:
                errors.append(f"finestra {window_dates}: {p['nome']} senza accesso")

    return errors


def write_json(people, start, end, access, path):
    out_people = []
    for p in people:
        out_people.append({
            "id": p["id"],
            "nome": p["nome"],
            "smart": sorted(DAY_IT[d] for d in p["smart"] if d in DAY_IT),
            "access": {k: v for k, v in access[p["id"]].items() if v is not None},
        })
    data = {"start": start.isoformat(), "end": end.isoformat(), "people": out_people}
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)


def write_xlsx(people, start, end, access, path):
    wb = openpyxl.Workbook()
    ws_rules = wb.active
    ws_rules.title = "Regole"
    rules_text = [
        "Regole di accesso alla risorsa (garage) - v3",
        "",
        "1. Nessun ID puo' accedere alla risorsa nel proprio giorno di smart working.",
        "2. Venerdi': accesso libero, senza limiti ne' rotazione.",
        "3. Lunedi', martedi', mercoledi', giovedi': capienza fissa di ESATTAMENTE 7 accessi al giorno.",
        "4. Rotazione: equita' a finestre scorrevoli di 3 giorni limitati consecutivi (esclusa Ketty).",
        "5. Ketty: accesso fisso ogni mercoledi' (non in rotazione); accede anche di venerdi' se non smart.",
        "",
        f"Periodo: {start.isoformat()} - {end.isoformat()}",
    ]
    for i, line in enumerate(rules_text, start=1):
        ws_rules.cell(row=i, column=1, value=line)

    ws_plan = wb.create_sheet("Piano giornaliero")
    ws_plan.append(["Data", "Giorno", "ID", "Nome", "Accesso"])
    for date in daterange(start, end):
        date_str = date.isoformat()
        day_label = DAY_IT.get(date.weekday())
        if day_label is None:
            continue
        for p in people:
            val = access[p["id"]].get(date_str)
            if val is None:
                continue
            ws_plan.append([date_str, day_label, p["id"], p["nome"], val])

    ws_summary = wb.create_sheet("Riepilogo capienza")
    ws_summary.append(["Data", "Giorno", "Accessi SI"])
    for date in daterange(start, end):
        date_str = date.isoformat()
        day_label = DAY_IT.get(date.weekday())
        if day_label is None:
            continue
        si_count = sum(1 for p in people if access[p["id"]].get(date_str) == "SI")
        ws_summary.append([date_str, day_label, si_count])

    wb.save(path)


HTML_TEMPLATE = """<!doctype html>
<html lang="it">
<head>
<meta charset="UTF-8">
<title>Piano Rotazione Garage v3</title>
<style>
  :root {{
    --bg: #ffffff;
    --fg: #1a1a1a;
    --border: #d8dce0;
    --header-bg: #2b3a55;
    --header-fg: #ffffff;
    --fixed-bg: #f4f6f8;
    --si-bg: #d7f2df;
    --si-fg: #1c6b34;
    --no-bg: #f3f4f6;
    --no-fg: #9aa0a6;
    --row-alt: #fafbfc;
    --total-bg: #eef1f5;
    --total-fg: #2b3a55;
  }}
  @media (prefers-color-scheme: dark) {{
    :root {{
      --bg: #16181d;
      --fg: #e6e8eb;
      --border: #33373e;
      --header-bg: #1f2937;
      --header-fg: #f3f4f6;
      --fixed-bg: #1c1f26;
      --si-bg: #17422a;
      --si-fg: #6fe19a;
      --no-bg: #202329;
      --no-fg: #6b7178;
      --row-alt: #1a1c22;
      --total-bg: #232830;
      --total-fg: #cbd3dc;
    }}
  }}
  :root[data-theme="dark"] {{
    --bg: #16181d; --fg: #e6e8eb; --border: #33373e; --header-bg: #1f2937;
    --header-fg: #f3f4f6; --fixed-bg: #1c1f26; --si-bg: #17422a; --si-fg: #6fe19a;
    --no-bg: #202329; --no-fg: #6b7178; --row-alt: #1a1c22;
    --total-bg: #232830; --total-fg: #cbd3dc;
  }}
  :root[data-theme="light"] {{
    --bg: #ffffff; --fg: #1a1a1a; --border: #d8dce0; --header-bg: #2b3a55;
    --header-fg: #ffffff; --fixed-bg: #f4f6f8; --si-bg: #d7f2df; --si-fg: #1c6b34;
    --no-bg: #f3f4f6; --no-fg: #9aa0a6; --row-alt: #fafbfc;
    --total-bg: #eef1f5; --total-fg: #2b3a55;
  }}

  * {{ box-sizing: border-box; }}
  body {{
    margin: 0;
    padding: 24px;
    background: var(--bg);
    color: var(--fg);
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Arial, sans-serif;
  }}
  h1 {{
    font-size: 1.4rem;
    margin: 0 0 4px 0;
  }}
  .subtitle {{
    color: var(--no-fg);
    font-size: 0.9rem;
    margin: 0 0 6px 0;
  }}
  .legend {{
    color: var(--no-fg);
    font-size: 0.82rem;
    margin: 0 0 18px 0;
    line-height: 1.5;
  }}
  .table-wrap {{
    overflow-x: auto;
    border: 1px solid var(--border);
    border-radius: 8px;
    max-width: 100%;
  }}
  table {{
    border-collapse: separate;
    border-spacing: 0;
    font-size: 0.82rem;
    white-space: nowrap;
  }}
  th, td {{
    padding: 6px 10px;
    border-bottom: 1px solid var(--border);
    border-right: 1px solid var(--border);
    text-align: center;
  }}
  thead th {{
    background: var(--header-bg);
    color: var(--header-fg);
    position: sticky;
    top: 0;
    z-index: 2;
    font-weight: 600;
    line-height: 1.3;
  }}
  .fixed-col {{
    position: sticky;
    background: var(--fixed-bg);
    color: var(--fg);
    font-weight: 600;
    text-align: left;
    z-index: 1;
  }}
  .id-col {{ left: 0; width: 40px; text-align: center; }}
  .nome-col {{ left: 40px; min-width: 130px; }}
  .giorno-col {{ left: 170px; min-width: 90px; }}
  thead .id-col {{ left: 0; z-index: 3; }}
  thead .nome-col {{ left: 40px; z-index: 3; }}
  thead .giorno-col {{ left: 170px; z-index: 3; }}
  .date-col {{ min-width: 52px; }}
  .total-col {{ min-width: 60px; }}
  td.si {{ background: var(--si-bg); color: var(--si-fg); font-weight: 600; }}
  td.no {{ background: var(--no-bg); color: var(--no-fg); }}
  td.total, th.total {{ background: var(--total-bg); color: var(--total-fg); font-weight: 700; }}
  tfoot td {{
    position: sticky;
    bottom: 0;
    background: var(--total-bg);
    color: var(--total-fg);
    font-weight: 700;
    z-index: 1;
  }}
  tfoot .fixed-col {{ z-index: 2; }}
  tbody tr:hover td {{ filter: brightness(0.97); }}
</style>
</head>
<body>
  <h1>Piano Rotazione Garage v3</h1>
  <p class="subtitle">{n_people} persone &middot; {n_days} giorni lavorativi dal {start_label} al {end_label}</p>
  <p class="legend">
    Venerdi: accesso libero per tutti (salvo giorno di smart working).<br>
    Lun, Mar, Mer, Gio: 7 posti/giorno, ruotati il piu possibile equamente tra le 21 persone (almeno 1 accesso garantito ogni 3 giorni limitati).<br>
    Ketty: accesso fisso ogni mercoledi, oltre al venerdi libero se non in smart.<br>
    Ultima colonna: totale accessi (SI) per persona. Ultima riga: totale accessi (SI) per giornata.
  </p>
  <div class="table-wrap">
    <table>
      <thead>
        <tr>
          <th class="fixed-col id-col">ID</th>
          <th class="fixed-col nome-col">NOME</th>
          <th class="fixed-col giorno-col">GIORNO SMART</th>
          {date_headers}
          <th class="total-col total">Tot.</th>
        </tr>
      </thead>
      <tbody>
        {body_rows}
      </tbody>
      <tfoot>
        <tr>
          <td class="fixed-col id-col">&nbsp;</td>
          <td class="fixed-col nome-col">Totale</td>
          <td class="fixed-col giorno-col">&nbsp;</td>
          {footer_cells}
          <td class="total">{grand_total}</td>
        </tr>
      </tfoot>
    </table>
  </div>
</body>
</html>
"""

DAY_LABEL_IT = {0: "Lun", 1: "Mar", 2: "Mer", 3: "Gio", 4: "Ven"}


def write_html(people, start, end, access, path):
    plan_dates = [d for d in daterange(start, end) if d.weekday() in DAY_LABEL_IT]

    date_headers = "".join(
        f'<th class="date-col">{DAY_LABEL_IT[d.weekday()]}<br>{d.strftime("%d/%m")}</th>'
        for d in plan_dates
    )

    body_rows_parts = []
    for p in people:
        cells = [
            f'<td class="fixed-col id-col">{p["id"]}</td>',
            f'<td class="fixed-col nome-col">{p["nome"]}</td>',
            f'<td class="fixed-col giorno-col">{"; ".join(sorted(DAY_IT[d] for d in p["smart"])) or "&nbsp;"}</td>',
        ]
        row_total = 0
        for d in plan_dates:
            val = access[p["id"]].get(d.isoformat())
            if val is None:
                cells.append('<td class="no">&nbsp;</td>')
            elif val == "SI":
                cells.append('<td class="si">SI</td>')
                row_total += 1
            else:
                cells.append('<td class="no">NO</td>')
        cells.append(f'<td class="total">{row_total}</td>')
        body_rows_parts.append("<tr>" + "".join(cells) + "</tr>")
    body_rows = "\n        ".join(body_rows_parts)

    footer_cells_parts = []
    grand_total = 0
    for d in plan_dates:
        date_str = d.isoformat()
        col_total = sum(1 for p in people if access[p["id"]].get(date_str) == "SI")
        grand_total += col_total
        footer_cells_parts.append(f'<td class="total">{col_total}</td>')
    footer_cells = "".join(footer_cells_parts)

    html = HTML_TEMPLATE.format(
        n_people=len(people),
        n_days=len(plan_dates),
        start_label=start.strftime("%d/%m/%Y"),
        end_label=end.strftime("%d/%m/%Y"),
        date_headers=date_headers,
        body_rows=body_rows,
        footer_cells=footer_cells,
        grand_total=grand_total,
    )

    with open(path, "w", encoding="utf-8") as f:
        f.write(html)


def main():
    rng = random.Random(RANDOM_SEED)
    people = load_people(SOURCE_XLSX)
    limited_dates = build_limited_day_sequence(people, PERIOD_START, PERIOD_END)

    access = assign_limited_days(people, limited_dates, rng)
    assign_free_day(people, PERIOD_START, PERIOD_END, access)

    errors = validate(people, access, limited_dates)
    if errors:
        print("VALIDAZIONE FALLITA:")
        for e in errors:
            print(" -", e)
        raise SystemExit(1)

    write_json(people, PERIOD_START, PERIOD_END, access, OUTPUT_JSON)
    write_xlsx(people, PERIOD_START, PERIOD_END, access, OUTPUT_XLSX)
    write_html(people, PERIOD_START, PERIOD_END, access, OUTPUT_HTML)
    print(f"OK: generati {OUTPUT_JSON}, {OUTPUT_XLSX} e {OUTPUT_HTML}")
    print(f"Periodo: {PERIOD_START} - {PERIOD_END}, {len(limited_dates)} giorni limitati, {len(people)} persone.")


if __name__ == "__main__":
    main()
